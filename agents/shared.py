import json
from datetime import datetime
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "Copie de banking_customers.xlsx"


def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH)


def sheet_to_records(ws) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            record = {}
            for header, value in zip(headers, row):
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                record[header] = value
            records.append(record)
    return records


def get_records(sheet_name: str, filter_key: str = None, filter_val: str = None) -> list[dict]:
    wb = load_workbook()
    ws = wb[sheet_name]
    records = sheet_to_records(ws)
    if filter_key and filter_val:
        records = [r for r in records if str(r.get(filter_key, "")) == filter_val]
    return records


def to_json(data) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)
