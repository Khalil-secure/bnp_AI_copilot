import json
import os
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from strands import Agent, tool

load_dotenv()

# Fix Windows terminal encoding so French accents display correctly
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


EXCEL_PATH = Path(__file__).parent / "Copie de banking_customers.xlsx"


def _load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH)


def _sheet_to_records(ws) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            record = {}
            for header, value in zip(headers, row):
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                record[header] = value
            records.append(record)
    return records


@tool
def list_sheets() -> str:
    """List all available sheets in the banking customers Excel file."""
    wb = _load_workbook()
    info = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        info.append({
            "sheet": name,
            "rows": ws.max_row - 1,
            "columns": headers,
        })
    return json.dumps(info, ensure_ascii=False, indent=2)


@tool
def get_sheet_data(sheet_name: str, client_id: str = None) -> str:
    """
    Extract all records from a sheet, optionally filtered by client_id.

    Args:
        sheet_name: Exact sheet name (e.g. '01_Profil_Client')
        client_id: Optional client ID to filter rows (e.g. 'CLI00001')
    """
    wb = _load_workbook()
    if sheet_name not in wb.sheetnames:
        return json.dumps({"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"})
    ws = wb[sheet_name]
    records = _sheet_to_records(ws)
    if client_id:
        records = [r for r in records if str(r.get("client_id", "")) == client_id]
    return json.dumps(records, ensure_ascii=False, indent=2, default=str)


@tool
def get_client_full_profile(client_id: str) -> str:
    """
    Extract all data for a single client across every sheet.

    Args:
        client_id: The client ID to retrieve (e.g. 'CLI00001')
    """
    wb = _load_workbook()
    profile = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records = _sheet_to_records(ws)
        matched = [r for r in records if str(r.get("client_id", "")) == client_id]
        if matched:
            profile[sheet_name] = matched
    if not profile:
        return json.dumps({"error": f"No data found for client '{client_id}'"})
    return json.dumps(profile, ensure_ascii=False, indent=2, default=str)


@tool
def list_clients() -> str:
    """List all client IDs and their names from the profile sheet."""
    wb = _load_workbook()
    ws = wb["01_Profil_Client"]
    records = _sheet_to_records(ws)
    clients = [
        {
            "client_id": r.get("client_id"),
            "nom": r.get("nom"),
            "prenom": r.get("prenom"),
            "archetype": r.get("archetype"),
        }
        for r in records
    ]
    return json.dumps(clients, ensure_ascii=False, indent=2)


agent = Agent(
    system_prompt=(
        "You are a banking data analyst assistant. "
        "You have access to an Excel file containing banking customer data across 11 sheets: "
        "client profiles, household info, projects, contracts, asset positions, valuation history, "
        "market indices, cash flows, and interaction events. "
        "Use your tools to extract and analyze the data as requested. "
        "Always respond in the same language the user writes in."
    ),
    tools=[list_sheets, get_sheet_data, get_client_full_profile, list_clients],
)


if __name__ == "__main__":
    print("Banking Customer Data Extraction Agent")
    print("=" * 40)
    print("Type 'exit' to quit.\n")
    while True:
        try:
            user_input = input("You: ").strip()
            if user_input.lower() in ("exit", "quit", "q"):
                break
            if not user_input:
                continue
            print()
            agent(user_input)
            print()
        except KeyboardInterrupt:
            print("\nGoodbye!")
            break
